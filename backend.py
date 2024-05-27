import openpyxl
import fitz
import os
import json
import datetime
folder = r"static/PDF"
list_db = []
from werkzeug.utils import secure_filename
def get_excel_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    max_row = sheet.max_row
    data_list = []
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        data_list.extend(list((row)))
    workbook.close()
    listt = [str(num) for num in data_list]
    return listt 

def PF(pdf):
    doc = fitz.open(pdf)
    MIC,RMW,SMW = [],[],[]
    mic = "BGBNG0016858000"
    rmw = "PYPNY2401211000"
    smw = "PYPNY2401591000"
    for i in range(len(doc)):
        page_no = doc[i]
        if page_no.search_for(mic):
            MIC.append(i+1)
        elif page_no.search_for(rmw):
            RMW.append(i+1)
        elif page_no.search_for(smw):
            SMW.append(i+1)
    return MIC[::len(MIC)-1],RMW[::len(RMW)-1],SMW[::len(SMW)-1]

def highlight_for_all_pf(excel,pdf,page,color=(0,0,1)):
    mic,rmw,smw = page
    mic_id = "BGBNG0016858000"
    rmw_id = "PYPNY2401211000"
    smw_id = "PYPNY2401591000"
    m,r,s = [],[],[]
    pageNoeach ={}
    page_start = 0
    doc = fitz.open(pdf)
    for uan in excel:
        for i in range(page_start,len(doc)):
            page_no = doc[i]
            text = page_no.search_for(uan,quads=True)
            MIC = page_no.search_for(mic_id)
            RMW = page_no.search_for(rmw_id)
            SMW = page_no.search_for(smw_id)
            if text and MIC:
                htext =  page_no.add_highlight_annot(text)
                htext.set_colors(stroke=color)
                htext.update(opacity=0.5)
                m.append(i+1)
                page_start = i
                break
            elif text and RMW:
                htext =  page_no.add_highlight_annot(text)
                htext.set_colors(stroke=color)
                htext.update(opacity=0.5)
                r.append(i+1)
                page_start = i
                break
            elif text and SMW:
                htext =  page_no.add_highlight_annot(text)
                htext.set_colors(stroke=color)
                htext.update(opacity=0.5)
                s.append(i+1)
                page_start = i
                break
    doc.save(pdf,incremental=True,encryption=0)

    doc.close()

    mic_page = sorted(set(m))
    rmw_page = sorted(set(r))
    smw_page = sorted(set(s))

    for index,pgeno in enumerate(mic_page):
        mic.insert(1+index,pgeno)
    for index,pgeno in enumerate(rmw_page):
        rmw.insert(1+index,pgeno)
    for index,pgeno in enumerate(smw_page):
        smw.insert(1+index,pgeno)

    for value ,key in zip(['MICRON','RMW','SMW'],[mic,rmw,smw]):
        pageNoeach[value] = ",".join([ str(pg) for pg in key])
    return pageNoeach

def only_micron_pf(excel,pdf,page,color=(0,0,1)):
    page = page[0]
    page_start = 0
    doc = fitz.open(pdf)
    page_num = []
    for uan in excel:
          for i in range(page_start,len(doc)):
            page_no = doc[i]
            text = page_no.search_for(uan,quads=True)
            if text:
                htext = page_no.add_highlight_annot(text)
                htext.set_colors(stroke=color)
                htext.update(opacity=0.5)
                page_num.append(i+1)
                page_start = i
                break
    doc.save(pdf,incremental=True,encryption=0)
    doc.close

    page_num = sorted(set(page_num))

    for pgno ,act_page in enumerate(page_num):
        page.insert(1+pgno,act_page)

    highlight_pageno = [ str(pg) for pg in page]
    page = {"Micron":",".join(highlight_pageno)}
    return page

def color_selecton(color):
    if color == 'red':
        return (1,0,0)
    elif color == 'green':
        return (0,0.9,0)
    elif color == 'blue':
        return (0,0,1)

def ESIC(pdf):
    doc = fitz.open(pdf)
    MIC = "53000116040000604"
    RMW = "53000563650001099"
    SMW = "53000563750001099"
    mic,rmw,smw = [],[],[]
    for pg in range(len(doc)):
        page_no = doc[pg]
        if page_no.search_for(MIC) :
            mic.append(pg+1) 
        elif page_no.search_for(RMW):
            mic.append((pg+1)-1)
            rmw.append(pg+1)
        elif  page_no.search_for(SMW):
            rmw.append((pg+1)-1)
            smw.append(pg+1)
    smw.append(len(doc))
    return mic,rmw,smw

def highlight_for_all_esic(excel,pdf,page,color=(0,0,1)):
    doc = fitz.open(pdf)
    mic,rmw,smw = page
    m,r,s = [],[],[]
    pageNoeach ={}
    page_start = 0
    for uan in excel:
        for pg in range(page_start,len(doc)):
            page_no = doc[pg]
            text = page_no.search_for(uan,quads=True)
            mics,mice = mic[0],mic[1]
            rmws,rmwe = rmw[0],rmw[1]
            smws,smwe = smw[0],smw[1]
            if text and (pg+1 >= mics and pg+1 <= mice) :
                htext =  page_no.add_highlight_annot(text)
                htext.set_colors(stroke=color)
                htext.update(opacity=0.5)
                m.append(pg+1)
                page_start = pg
                break   
                
            if text and (pg+1 >= rmws and pg+1 <= rmwe) :
                htext =  page_no.add_highlight_annot(text)
                htext.set_colors(stroke=color)
                htext.update(opacity=0.5)
                r.append(pg+1)
                page_start = pg
                break   
                
            if text and (pg+1 >= smws and pg+1 <= smwe) :
                htext =  page_no.add_highlight_annot(text)
                htext.set_colors(stroke=color)
                htext.update(opacity=0.5)
                s.append(pg+1)
                page_start = pg
                break   
  
    doc.save(pdf,incremental=True,encryption=0)
    doc.close()
    

    # insert highlight page number to the company wise 
    for index,pgeno in enumerate(m):
        mic.insert(1+index,pgeno)
    for index,pgeno in enumerate(r):
        rmw.insert(1+index,pgeno)
    for index,pgeno in enumerate(s):
        smw.insert(1+index,pgeno)
    
    # Remove the duplicate page no in list
    mic_page = sorted(set(mic))
    rmw_page = sorted(set(rmw))
    smw_page = sorted(set(smw))
    # Remove the last page_number from the list
    mic_page.pop(len(mic_page)-1)
    rmw_page.pop(len(rmw_page)-1)
    smw_page.pop(len(smw_page)-1)
    # covert list of page into the company wise 
    for value ,key in zip(['MICRON','RMW','SMW'],[mic_page,rmw_page,smw_page]):
        pageNoeach[value] = ",".join([ str(pg) for pg in key])

    return pageNoeach

def only_micron_esi(excel,pdf,page,color=(0,0,1)):
    page = page[0]
    page_start = 0
    doc = fitz.open(pdf)
    page_num = []
    for uan in excel:
          for i in range(page_start,len(doc)):
            page_no = doc[i]
            text = page_no.search_for(uan,quads=True)
            if text:
                htext = page_no.add_highlight_annot(text)
                htext.set_colors(stroke=color)
                htext.update(opacity=0.5)
                page_num.append(i+1)
                page_start = i
                break
    doc.save(pdf,incremental=True,encryption=0)
    doc.close

    for pgno ,act_page in enumerate(page_num):
        page.insert(1+pgno,act_page)
    
    mic_page = sorted(set(page))
    
    mic_page.pop(len(mic_page)-1)

    highlight_pageno = [ str(pg) for pg in mic_page]
    page = {"Micron": ",".join(highlight_pageno)}
    return page

def highlight_only(excel,pdf,color=(0,0,1)):
    page_start = 0
    doc = fitz.open(pdf)
    page_num = [1]
    for uan in excel:
          for i in range(page_start,len(doc)):
            page_no = doc[i]
            text = page_no.search_for(uan,quads=True)
            if text:
                htext = page_no.add_highlight_annot(text)
                htext.set_colors(stroke=color)
                htext.update(opacity=0.5)
                page_num.append(i+1)
                page_start = i
                break
    doc.save(pdf,incremental=True,encryption=0)
    doc.close
    page_num = sorted(set(page_num))
    highlight_pageno = [ str(pg) for pg in page_num]
    return ",".join(highlight_pageno)

def Change_filename(file,project_name):
    file_name = file.replace(" ", "_")
    return f"{project_name}_{file_name}"

def save_file(file,project_name):
    photo_name = secure_filename(Change_filename(file.filename, project_name))
    file.save(os.path.join(folder, photo_name))
